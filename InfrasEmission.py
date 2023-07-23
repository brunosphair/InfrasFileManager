import os
import re
import sys
from pathlib import Path
from easygui import buttonbox, ccbox, multchoicebox, enterbox, msgbox,\
                    multenterbox
from zipfile import ZipFile
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import datetime
from dotenv import dotenv_values


class Emission:
    def __init__(self):
        self.doc_reg_expression, self.rev_reg_expression = \
                                                    self.get_reg_expressions()
        self.issued_path = self.get_issued_path()
        self.docs = self.get_files()
        self.directories = self.get_directories()
        self.ld_number = self.get_ld_number()
        self.project_number = self.get_project_number()
        self.grd_number = self.get_grd_number()
        self.grd_name = 'IFS-GRD-' + \
                        str(self.project_number) + \
                        "-" + str(self.grd_number).zfill(3)
        self.file_num_caract = 23
        self.ld_information = {}

    def get_files(self):
        docs = []
        for file in os.listdir('.'):
            if os.path.isfile(file):
                rev = self.get_revision(file)
                docs.append([file, rev, True])
        return docs

    def get_issued_path(self):
        path = Path(os.getcwd()).parent.absolute()
        parent_path = path.parent.absolute()
        issued_path = os.path.join(parent_path, '3_Emitidos')
        if not os.path.isdir(issued_path):
            raise FileNotFoundError("A pasta 3_Emitidos não foi encontrada")
        return issued_path

    def get_directories(self):
        directories = []
        for file in os.listdir(self.issued_path):
            if os.path.isdir(os.path.join(self.issued_path, file)):
                directories.append(file)
        return directories

    def get_ld_number(self):
        lds_directory = os.path.join(self.issued_path, '_LDs')
        if not os.path.isdir(lds_directory):
            raise FileNotFoundError("A pasta _LDs não foi encontrada")
        lds = os.listdir(lds_directory)
        # print(lds)
        self.ld_name = 'IFS-XXXX-XXX-X-LD-XXXX.xlsx'
        last_revision = -1
        for item in lds:
            ld_revision = self.get_ld_revision(item)
            if last_revision < ld_revision:
                last_revision = ld_revision
                self.ld_name = item

        return last_revision

    def get_grd_number(self):
        book_path = os.path.join(self.issued_path, '_LDs', self.ld_name)
        wb = openpyxl.load_workbook(book_path, read_only=True)
        grd_number = 1
        sheet_name = 'GRD-' + str(grd_number).zfill(3)
        while sheet_name in wb.sheetnames:
            grd_number += 1
            sheet_name = 'GRD-' + str(grd_number).zfill(3)
        wb.close()
        # print(grd_number)
        return grd_number

    def get_reg_expressions(self):
        try:
            config = dotenv_values('config.env')
            print(config)
            doc_reg_expression = config['DOC_REGULAR_EXPRESSION']
            rev_reg_expression = config['REV_REGULAR_EXPRESSION']
        except KeyError:
            doc_reg_expression = \
                        r'^IFS-\d{4}-\d{3}-\w{1}-\w{2}-\d{5}.*(_R\d{1,2})?$'
            rev_reg_expression = r'(?i)_R\d+$'

        return doc_reg_expression, rev_reg_expression

    def get_project_number(self):
        path = Path(os.getcwd()).parent.absolute()
        project_path = path.parent.absolute()
        dir_name = os.path.basename(project_path)
        project_number = dir_name[:4]
        if not project_number.isnumeric():
            raise ValueError("O arquivo executado não está na pasta correta")
        return project_number

    def check_pattern(self):
        ignored_files = []
        for doc in self.docs:
            if not self.verify_pattern(doc[0]):
                ignored_files.append(doc[0])
                doc[2] = False
        if len(ignored_files):
            msg = "Os seguintes arquivos não serão emitidos, pois não "\
                "correspondem ao padrão de nomenclatura de arquivos:\n\n" \
                + '\n'.join(ignored_files) + "\n\nO que deseja fazer?"
            title = "Inconsistência na nomenclatura dos arquivos"
            if ccbox(msg, title):
                pass
            else:
                sys.exit(0)

    def check_files(self):
        for doc in self.docs:
            if doc[2]:
                path_name = doc[0][:self.file_num_caract]
                doc_name = self.get_file_name(doc[0])
                doc_directory = os.path.join(self.issued_path, path_name)
                if os.path.isdir(doc_directory):
                    for file in os.listdir(doc_directory):
                        try:
                            file_name = self.get_file_name(file)
                            if self.get_revision(file
                                                 ) == doc[1] and file_name == doc_name:
                                raise NameError
                        except NameError:
                            msg = 'O arquivo ' + doc_name + ' com a revisão '\
                                + str(doc[1]) \
                                + ' já existe. O que deseja fazer?'
                            choices = [
                                       "Não emitir esse arquivo",
                                       "Emitir mesmo assim",
                                       "Cancelar"
                                       ]
                            title = "Arquivo duplicado"
                            choice = buttonbox(msg, title, choices)
                            if choice == "Não emitir esse arquivo":
                                # print("arquivo ignorado")
                                doc[2] = False
                            elif choice == "Emitir mesmo assim":
                                obsolete_path = os.path.join(doc_directory,
                                                             "Obsoleto")
                                if not os.path.isdir(obsolete_path):
                                    os.mkdir(obsolete_path)
                                i = 1
                                file_aux = file
                                while os.path.isfile(
                                    os.path.join(obsolete_path,
                                                 file_aux)):
                                    file_aux = os.path.splitext(file)[0]\
                                        + "(" + str(i) + ")"\
                                        + os.path.splitext(file)[0]
                                    i += 1
                                file_source_path = os.path.join(doc_directory,
                                                                file)
                                file_destiny_path = os.path.join(obsolete_path,
                                                                 file_aux)
                                os.replace(file_source_path, file_destiny_path)
                                doc[2] = True
                            elif choice == "Cancelar":
                                sys.exit(0)
        list_of_options = []
        for doc in self.docs:
            if doc[2]:
                list_of_options.append(doc[0])
        if len(list_of_options) == 0:
            sys.exit(0)
        elif len(list_of_options) == 1:
            ccbox("O seguinte arquivo será emitido:\n\n" + list_of_options[0])
        else:
            msg = "Os seguintes arquivos serão emitidos na "\
                + self.grd_name\
                + ". Desmarque caso não queira emitir algum."
            title = 'Deseja continuar?'
            choices = multchoicebox(msg,
                                    title,
                                    list_of_options,
                                    preselect=[*range(len(list_of_options))])
            # print(choices)
            for doc in self.docs:
                if not doc[0] in choices:
                    doc[2] = False

    def create_zip(self):
        zipObj = ZipFile(self.grd_name + '.zip', 'w')
        for doc in self.docs:
            if doc[2]:
                zipObj.write(doc[0])
        zipObj.close()

    def create_ld(self):
        no_docs = []
        grd_items = []
        for doc in self.docs:
            doc_name = self.get_file_name(doc[0])
            if doc[2] and doc_name not in no_docs:
                no_docs.append(doc_name)
                grd_items.append([doc_name, doc[1]])
        self.create_excel_grd(grd_items)

    def get_ld_information(self):
        date_defined = False
        while not date_defined:
            title = "Data de emissão"
            text = "Digite a data de emissão da GRD no formato DD/MM/YY:"
            today_date = datetime.datetime.now().strftime("%d/%m/%y")
            emission_date = enterbox(text, title, today_date)
            if self.verify_date_pattern(emission_date):
                date_defined = True
            else:
                msgbox("Formato de data inválido")

        self.ld_information["emission_date"] = emission_date

        if self.ld_number == -1:
            text = "Como essa é a primeira emissão desse projeto, digite um "\
                "nome para a LD no padrão IFS-NNNN-NNN-X-LD-NNNNN onde X são "\
                "letras e N são números"
            title = "Nomeie a LD"
            probably_name = self.get_probably_name()
            d_text = "IFS-"\
                     + str(self.project_number)\
                     + "-" + probably_name + "-G-LD-00001"
            defined_name = False
            while not defined_name:
                ld_name = enterbox(text, title, d_text)
                if self.verify_ld_pattern_no_rev(ld_name):
                    defined_name = True
                else:
                    msgbox("O nome que você digitou não atende aos requisitos"
                           " de IFS-NNNN-NNN-X-LD-NNNNN, digite novamente",
                           "Nome inválido!")
            self.ld_information["ld_name"] = ld_name

            text = "Defina os títulos da LD:"
            title = "Definir títulos"
            input_list = ["1ª LINHA - TIPO DE PROJETO ",
                          "2ª LINHA - TÍTULO DO PROJETO",
                          "3ª LINHA - TÍTULO DO DOCUMENTO"]
            default_list = ["PROJETO CONCEITUAL/BÁSICO/EXECUTIVO",
                            "TÍTULO DO PROJETO",
                            "TÍTULO DO DOCUMENTO"]
            output = multenterbox(text, title, input_list, default_list)
            self.ld_information["project_title"] = output[1]
            self.ld_information["ld_title"] = "\n".join(output)\
                                              + "\nLISTA DE DOCUMENTOS"

        text = "Defina as iniciais dos responsáveis (formato XXX)"
        title = "Defina as iniciais"
        input_list = ["EXECUÇÃO", "VERIFICAÇÃO", "APROVAÇÃO"]
        if self.ld_number == -1:
            default_list = ["XXX", "XXX", "XXX"]
        else:
            revision = self.ld_number + 1
            previous_cover_cell = self.get_cover_cell(revision - 1)
            book_path = os.path.join(self.issued_path, '_LDs', self.ld_name)
            wb = openpyxl.load_workbook(book_path, read_only=True)
            cover_sheet = wb['Capa']
            d1 = cover_sheet.cell(row=previous_cover_cell[0] + 1,
                                  column=previous_cover_cell[1]).value
            d2 = cover_sheet.cell(row=previous_cover_cell[0] + 2,
                                  column=previous_cover_cell[1]).value
            d3 = cover_sheet.cell(row=previous_cover_cell[0] + 3,
                                  column=previous_cover_cell[1]).value
            default_list = [d1, d2, d3]
        output = multenterbox(text, title, input_list, default_list)
        self.ld_information["acronym1"] = output[0]
        self.ld_information["acronym2"] = output[1]
        self.ld_information["acronym3"] = output[2]

    def create_excel_grd(self, grd_items):
        book_path = os.path.join(self.issued_path, '_LDs', self.ld_name)
        book = openpyxl.load_workbook(book_path)
        template_sheet = book['GRD-XXX']
        cover_sheet = book['Capa']
        sheet = book.copy_worksheet(template_sheet)
        sheet.title = 'GRD-' + str(self.grd_number).zfill(3)
        i = 1
        for item in grd_items:
            sheet.cell(row=25 + i, column=1).value = int(i)
            sheet.cell(row=25 + i, column=2).value = item[0]
            sheet.cell(row=25 + i, column=16).value = int(item[1])
            i += 1
        sheet.cell(row=10, column=12).value = self.grd_name

        sheet.cell(row=7,
                   column=12).value = self.ld_information["emission_date"]

        yellowFill = PatternFill(start_color='FFFF00',
                                 end_color='FFFF00',
                                 fill_type='solid')
        sheet.conditional_formatting.add('$E$26:$O$192',
                                         FormulaRule(formula=[
                                                    'AND($B26<>"",E26="")'
                                                    ],
                                                    stopIfTrue=False,
                                                    fill=yellowFill
                                                    )
                                         )
        sheet.conditional_formatting.add('$Q$26:$R$192',
                                         FormulaRule(formula=[
                                                     'AND($B26<>"",Q26="")'
                                                     ],
                                                     stopIfTrue=False,
                                                     fill=yellowFill
                                                     )
                                         )

        if self.ld_number == -1:
            revision = 0
            ld_name = self.ld_information["ld_name"]
            cover_sheet.cell(row=2,
                             column=4).value = ld_name
            ld_name = ld_name + "_R0"
            project_title = self.ld_information["project_title"]
            cover_sheet.cell(row=5,
                             column=1).value = self.ld_information["ld_title"]
        else:
            revision = self.ld_number + 1
            ld_name = self.ld_name[:self.file_num_caract] \
                + '_R' \
                + str(revision)
            last_grd = book['GRD-' + str(self.grd_number - 1).zfill(3)]
            project_title = last_grd.cell(row=1, column=6).value
        sheet.cell(row=1, column=6).value = project_title
        cover_sheet.cell(row=6, column=12).value = revision
        cover_sheet.cell(row=16 + revision, column=1).value = revision
        cover_sheet.cell(row=16 + revision, column=2).value = "C"
        cover_sheet.cell(row=16 + revision, column=3).value = self.grd_name
        rev_cell = self.get_cover_cell(revision)
        rev_row = rev_cell[0]
        rev_column = rev_cell[1]

        cover_sheet.cell(row=rev_row,
                         column=rev_column).value = self.ld_information[
            "emission_date"]
        cover_sheet.cell(row=rev_row + 1,
                         column=rev_column).value = self.ld_information[
            "acronym1"]
        cover_sheet.cell(row=rev_row + 2,
                         column=rev_column).value = self.ld_information[
            "acronym2"]
        cover_sheet.cell(row=rev_row + 3,
                         column=rev_column).value = self.ld_information[
            "acronym3"]

        ld_final_path = os.path.join(self.issued_path,
                                     '_LDs',
                                     ld_name + '.xlsx')
        book.save(filename=ld_final_path)
        book.close()

    def check_open_files(self):
        file_open = True
        while file_open:
            try:
                for doc in self.docs:
                    if doc[2]:
                        src = Path(doc[0])
                        os.replace(src, src)
                file_open = False
            except OSError:
                file_open = True
                text = "O arquivo " + doc[0] + " está aberto. Feche-o e clique em repetir para continuar a operação."
                title = "Todos os arquivos devem estar fechados"
                button_list = ["Repetir", "Cancelar"]
                output = buttonbox(text, title, button_list)
                if output == "Repetir":
                    pass
                elif output == "Cancelar":
                    sys.exit(0)

    def move_files(self):
        # Deletes the revision suffix from the filename
        filenames = []
        for doc in self.docs:
            if doc[2]:
                filenames.append(doc[0][:self.file_num_caract])
        set(filenames)
        # If the file directory doesn't exists, the code creates it
        for item in filenames:
            if item not in self.directories:
                dir_to_create = os.path.join(self.issued_path, item)
                os.mkdir(dir_to_create)
                self.directories.append(item)
        for directory in self.directories:
            for doc in self.docs:
                if doc[2] and doc[0].startswith(directory):
                    src = Path(doc[0])
                    dest = Path(os.path.join(os.path.join(self.issued_path,
                                                          directory),
                                             doc[0]))
                    os.replace(src, dest)

    def get_file_name(self, doc):
        filename = doc[:self.file_num_caract]
        return filename

    def get_probably_name(self):
        doc_name = self.docs[0]
        probably_name = doc_name[0][9:12]

        return probably_name

    def get_revision(self, doc):
        filename = os.path.splitext(doc)[0]
        pattern = self.rev_reg_expression
        if re.search(pattern, os.path.splitext(filename)[0]) is not None:
            rev = re.search(pattern, os.path.splitext(filename)[0]).group()
            rev = int(''.join(filter(str.isdigit, rev)))
        else:
            rev = 0
        return rev

    def verify_pattern(self, doc_name):
        doc_name_no_extension = os.path.splitext(doc_name)[0]
        pattern = self.doc_reg_expression
        if re.match(pattern, doc_name_no_extension):
            return True
        else:
            return False

    @staticmethod
    def verify_ld_pattern_no_rev(doc_name):
        pattern = r'^IFS-\d{4}-\d{3}-\w{1}-LD-\d{5}$'
        if re.match(pattern, doc_name):
            return True
        else:
            return False

    @staticmethod
    def get_ld_revision(doc_name):
        doc_name_no_extension = os.path.splitext(doc_name)[0]
        pattern = r'^IFS-\d{4}-\d{3}-\w{1}-LD-\d{5}.*(_R\d{1,2})?$'
        if not re.match(pattern, doc_name_no_extension):
            return -1
        else:
            revision = re.search(r'(?i)_R\d+$',
                                 os.path.splitext(doc_name_no_extension)[0]
                                 ).group()
            revision = int(''.join(filter(str.isdigit, revision)))
            return revision

    @staticmethod
    def verify_date_pattern(date):
        pattern = r"\d{2}/\d{2}/\d{2}"
        if re.match(pattern, date):
            return True
        else:
            return False

    @staticmethod
    def get_cover_cell(rev):
        if rev == 0 or rev == 5:
            column = 3
        elif rev == 1 or rev == 6:
            column = 5
        elif rev == 2 or rev == 7:
            column = 7
        elif rev == 3 or rev == 8:
            column = 8
        elif rev == 4 or rev == 9:
            column = 11

        if rev <= 4:
            row = 32
        else:
            row = 37

        return [row, column]
    
    #TODO : Verificar arquivos que terminam com Rev


if __name__ == '__main__':
    os.chdir(r'C:\Users\Bruno\OneDrive\Documentos\LD\2227 Exemplo\5_Engenharia\_PARA EMISSAO')
    emis = Emission()
    emis.check_pattern()
    emis.check_files()
    emis.get_ld_information()
    emis.check_open_files()
    emis.create_zip()
    emis.create_ld()
    emis.move_files()
    # print(emis.get_project_number())
    # print(emis.verify_pattern("IFS-2122-110-B-CP-00001_R0"))
