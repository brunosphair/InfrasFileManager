import openpyxl
import os

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


def get_grd_number(emited_path, ld_name):
    '''
    Opens the excel LD and returns the number of the GRD that are going to be
    issued
    '''
    book_path = os.path.join(emited_path, '_LDs', ld_name)
    wb = openpyxl.load_workbook(book_path, read_only=True)
    grd_number = 1
    sheet_name = 'GRD-' + str(grd_number).zfill(3)
    while sheet_name in wb.sheetnames:
        grd_number += 1
        sheet_name = 'GRD-' + str(grd_number).zfill(3)
    wb.close()

    return grd_number


def create_excel_grd(emited_path, ld_name, grd_number, grd_name,
                     ld_information, ld_rev, file_num_caract, grd_items):
    book_path = os.path.join(emited_path, '_LDs', ld_name)
    book = openpyxl.load_workbook(book_path)
    template_sheet = book['GRD-XXX']
    cover_sheet = book['Capa']
    sheet = book.copy_worksheet(template_sheet)
    sheet.title = 'GRD-' + str(grd_number).zfill(3)
    i = 1
    for item in grd_items:
        sheet.cell(row=25 + i, column=1).value = int(i)
        sheet.cell(row=25 + i, column=2).value = item[0]
        sheet.cell(row=25 + i, column=16).value = int(item[1])
        i += 1
    sheet.cell(row=10, column=12).value = grd_name

    sheet.cell(row=7,
               column=12).value = ld_information["emission_date"]

    yellowFill = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')
    sheet.conditional_formatting.add('$E$26:$O$192',
                                     FormulaRule(formula=[
                                                 'AND($B26<>"",E26="")'
                                                 ],
                                                 stopIfTrue=False,
                                                 fill=yellowFill
                                                 )
                                     )
    sheet.conditional_formatting.add('$Q$26:$R$192',
                                     FormulaRule(formula=[
                                                 'AND($B26<>"",Q26="")'
                                                 ],
                                                 stopIfTrue=False,
                                                 fill=yellowFill
                                                 )
                                     )

    if ld_rev == -1:
        revision = 0
        ld_name = ld_information["ld_name"]
        cover_sheet.cell(row=2, column=4).value = ld_name
        ld_name = ld_name + "_R0"
        project_title = ld_information["project_title"]
        cover_sheet.cell(row=5, column=1).value = ld_information["ld_title"]
    else:
        revision = ld_rev + 1
        ld_name = ld_name[:file_num_caract] \
            + '_R' \
            + str(revision)
        last_grd = book['GRD-' + str(grd_number - 1).zfill(3)]
        project_title = last_grd.cell(row=1, column=6).value
    sheet.cell(row=1, column=6).value = project_title
    cover_sheet.cell(row=6, column=12).value = revision
    if revision > 13:
        reorder_description_cells(cover_sheet)
    cover_sheet.cell(row=16 + revision, column=1).value = revision
    cover_sheet.cell(row=16 + revision, column=2).value = "C"
    cover_sheet.cell(row=16 + revision, column=3).value = grd_name
    rev_cell = get_cover_cell(revision)
    rev_row = rev_cell[0]
    rev_column = rev_cell[1]

    if revision > 9:
        reorder_rev_cells(cover_sheet, revision)

    cover_sheet.cell(row=rev_row,
                     column=rev_column).value = ld_information[
        "emission_date"]
    cover_sheet.cell(row=rev_row + 1,
                     column=rev_column).value = ld_information["acronym1"]
    cover_sheet.cell(row=rev_row + 2,
                     column=rev_column).value = ld_information["acronym2"]
    cover_sheet.cell(row=rev_row + 3,
                     column=rev_column).value = ld_information["acronym3"]

    ld_final_path = os.path.join(emited_path,
                                 '_LDs',
                                 ld_name + '.xlsx')
    book.save(filename=ld_final_path)
    book.close()


def get_cover_cell(rev):
    if rev == 0 or rev == 5:
        column = 3
    elif rev == 1 or rev == 6:
        column = 5
    elif rev == 2 or rev == 7:
        column = 7
    elif rev == 3 or rev == 8:
        column = 8
    elif rev == 4 or rev == 9:
        column = 11

    if rev <= 4:
        row = 32
    else:
        row = 37

    return [row, column]


def get_acronym_default_list(book_path, previous_cover_cell):
    wb = openpyxl.load_workbook(book_path, read_only=True)
    cover_sheet = wb['Capa']
    d1 = cover_sheet.cell(row=previous_cover_cell[0] + 1,
                          column=previous_cover_cell[1]).value
    d2 = cover_sheet.cell(row=previous_cover_cell[0] + 2,
                          column=previous_cover_cell[1]).value
    d3 = cover_sheet.cell(row=previous_cover_cell[0] + 3,
                          column=previous_cover_cell[1]).value
    return [d1, d2, d3]


def reorder_rev_cells(cover_sheet, revision):

    # Column G to E - LINHA 1
    copy_values(cover_sheet, 31, 7, 31, 5)
    copy_values(cover_sheet, 32, 7, 32, 5)
    copy_values(cover_sheet, 33, 7, 33, 5)
    copy_values(cover_sheet, 34, 7, 34, 5)
    copy_values(cover_sheet, 35, 7, 35, 5)
    # Column H to G LINHA 1
    copy_values(cover_sheet, 31, 8, 31, 7)
    copy_values(cover_sheet, 32, 8, 32, 7)
    copy_values(cover_sheet, 33, 8, 33, 7)
    copy_values(cover_sheet, 34, 8, 34, 7)
    copy_values(cover_sheet, 35, 8, 35, 7)
    # Column K to H - LINHA 1
    copy_values(cover_sheet, 31, 11, 31, 8)
    copy_values(cover_sheet, 32, 11, 32, 8)
    copy_values(cover_sheet, 33, 11, 33, 8)
    copy_values(cover_sheet, 34, 11, 34, 8)
    copy_values(cover_sheet, 35, 11, 35, 8)
    # Column C to K - LINHA 2 P/ 1
    copy_values(cover_sheet, 36, 3, 31, 11)
    copy_values(cover_sheet, 37, 3, 32, 11)
    copy_values(cover_sheet, 38, 3, 33, 11)
    copy_values(cover_sheet, 39, 3, 34, 11)
    copy_values(cover_sheet, 40, 3, 35, 11)
    # Column E to C - LINHA 2
    copy_values(cover_sheet, 36, 5, 36, 3)
    copy_values(cover_sheet, 37, 5, 37, 3)
    copy_values(cover_sheet, 38, 5, 38, 3)
    copy_values(cover_sheet, 39, 5, 39, 3)
    copy_values(cover_sheet, 40, 5, 40, 3)
    # Column G to E - LINHA 2
    copy_values(cover_sheet, 36, 7, 36, 5)
    copy_values(cover_sheet, 37, 7, 37, 5)
    copy_values(cover_sheet, 38, 7, 38, 5)
    copy_values(cover_sheet, 39, 7, 39, 5)
    copy_values(cover_sheet, 40, 7, 40, 5)
    # Column G to E - LINHA 2
    copy_values(cover_sheet, 36, 7, 36, 5)
    copy_values(cover_sheet, 37, 7, 37, 5)
    copy_values(cover_sheet, 38, 7, 38, 5)
    copy_values(cover_sheet, 39, 7, 39, 5)
    copy_values(cover_sheet, 40, 7, 40, 5)
    # Column H to G - LINHA 2
    copy_values(cover_sheet, 36, 7, 36, 5)
    copy_values(cover_sheet, 37, 8, 37, 7)
    copy_values(cover_sheet, 38, 8, 38, 7)
    copy_values(cover_sheet, 39, 8, 39, 7)
    copy_values(cover_sheet, 40, 8, 40, 7)
    # Column K to H - LINHA 2
    copy_values(cover_sheet, 36, 11, 36, 8)
    copy_values(cover_sheet, 37, 11, 37, 8)
    copy_values(cover_sheet, 38, 11, 38, 8)
    copy_values(cover_sheet, 39, 11, 39, 8)
    copy_values(cover_sheet, 40, 11, 40, 8)

    cover_sheet.cell(row=36, column=11).value = "REV. " + str(revision)


def reorder_description_cells(cover_sheet):
    for row in reversed(range(13)):
        copy_values(cover_sheet, row + 17, 1, row + 16, 1)
        copy_values(cover_sheet, row + 17, 2, row + 16, 2)
        copy_values(cover_sheet, row + 17, 3, row + 16, 3)

def copy_values(cover_sheet, from_row, from_column, to_row, to_column):
    cover_sheet.cell(row=to_row,
                     column=to_column).value = cover_sheet.cell(row=from_row, column=from_column).value

